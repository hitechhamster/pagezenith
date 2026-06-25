"""把每次分析结果追加到本地 Excel（一行一份报告）。

文件被 Excel 占用（用户正打开）时会写到带时间戳的副本，绝不让落盘失败影响接口。
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .config import get_settings
from .models import AnalysisReport

logger = logging.getLogger(__name__)

HEADERS = [
    "时间", "模式", "关键词", "目标URL", "进前10", "排名",
    "信息增益", "冗余度", "可读性(相对)", "Flesch", "前10Flesch均值",
    "经验", "专业", "可信", "权威", "权威可用",
    "缺失基本盘数", "独特内容数",
    "优先行动(Top5)", "缺失内容(Top10)",
]


def _row(r: AnalysisReport) -> list:
    ig, rd, ee = r.information_gain, r.readability, r.eeat
    verdict_zh = {"easier": "偏易", "aligned": "匹配", "harder": "偏难", "n/a": "—"}
    actions = "\n".join(f"[{a.impact}] {a.action}" for a in r.priority_actions[:5])
    missing = "\n".join(
        f"[{m.covered_by_n_competitors}家] {m.text}" for m in ig.missing_points[:10]
    )
    return [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "对比竞品" if r.mode == "keyword" else "单页",
        r.keyword or "",
        r.target_url,
        "是" if r.in_top_10 else "否",
        r.rank if r.rank is not None else "",
        ig.gain_score if ig.gain_score is not None else "",
        ig.redundancy_ratio,
        verdict_zh.get(rd.verdict, rd.verdict),
        rd.metrics.flesch_reading_ease,
        rd.top10_avg.flesch_reading_ease or "",
        ee.experience.score,
        ee.expertise.score,
        ee.trust.score,
        ee.authoritativeness.score if ee.authoritativeness.available else "",
        "是" if ee.authoritativeness.available else "否",
        len(ig.missing_points),
        len(ig.novel_points),
        actions,
        missing,
    ]


def append_report(report: AnalysisReport, path: str | None = None) -> str:
    """追加一行，返回实际写入的文件路径。失败抛异常由上层吞掉。"""
    p = Path(path or get_settings().excel_path)
    p.parent.mkdir(parents=True, exist_ok=True)

    target = p
    try:
        if target.exists():
            wb = load_workbook(target)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "SEO报告"
            ws.append(HEADERS)
        ws.append(_row(report))
        wb.save(target)
    except PermissionError:
        # 文件被 Excel 占用 → 写到带时间戳的副本
        target = p.with_name(f"{p.stem}_{datetime.now():%Y%m%d_%H%M%S}{p.suffix}")
        wb = Workbook()
        ws = wb.active
        ws.title = "SEO报告"
        ws.append(HEADERS)
        ws.append(_row(report))
        wb.save(target)
        logger.warning("Excel 主文件被占用，已写入副本 %s", target)
    return str(target)
