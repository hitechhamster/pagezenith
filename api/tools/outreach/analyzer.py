"""外链拓客编排：足迹找站 → 抓取分类 → 抓邮箱 → 可选生成邮件 → 流式产出 + Excel。"""

from __future__ import annotations

import asyncio
import io
import logging
from datetime import datetime
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from ..seo_gap.clients.fetch import PageFetcher, looks_blocked
from ..seo_gap.clients.llm import LLMClient
from ..seo_gap.clients.serpapi import SerpApiClient
from ..seo_gap.config import Settings, get_settings
from .emails import extract_emails, find_contact_links, has_contact_form
from .footprints import build_footprints
from .models import OutreachRequest, Prospect, ProspectEmail
from .prompts import CLASSIFY_SYSTEM, build_classify_user

logger = logging.getLogger(__name__)


def _domain(url: str) -> str:
    return urlparse(url).netloc.replace("www.", "").lower()


_CLS_MOCK = {"site_type": "博客", "relevance": 72, "opportunity": "投稿",
             "reason": "主题相关的独立博客，接受投稿。",
             "email_subject": "Guest post idea for your readers",
             "email_body": "Hi, I loved your recent piece on ... I'd like to contribute a "
                           "practical guide on ... Would you be open to a guest post? Thanks!"}


class OutreachFinder:
    def __init__(self, settings: Settings | None = None):
        self.s = settings or get_settings()
        self.serp = SerpApiClient(self.s)
        self.llm = LLMClient(self.s)
        self.fetcher = PageFetcher(self.s)

    async def _derive_keyword(self, url: str) -> str:
        """网址驱动、未给关键词时：抓页面让 LLM 提一个最合适的英文主题词。"""
        try:
            page = await self.fetcher.fetch(url)
        except Exception:
            return ""
        sample = ((page.title or "") + "\n" + page.text)[:2500]
        mock = {"keyword": "forex broker scam"}
        raw = await self.llm.complete_json(
            "从给定网页内容提炼一个最适合用来找外链机会的英文主题词/短语。只输出 JSON。",
            f"网页内容：\n{sample}\n\n输出 JSON：{{\"keyword\":\"...\"}}", mock=mock)
        return (raw.get("keyword", "") if isinstance(raw, dict) else "") or ""

    async def discover(self, keyword: str, loc: int, lang: str, your_domain: str,
                       cap: int) -> list[dict]:
        """跑足迹搜索，按域名去重，返回 [{domain,url,title,hint}]（最多 cap 个）。"""
        fps = build_footprints(keyword, self.s.outreach_footprints_per_run)
        results = await asyncio.gather(*[
            self.serp.fetch_serp(q, loc, lang, depth=10) for q, _ in fps
        ], return_exceptions=True)
        seen, out = set(), []
        for (q, hint), items in zip(fps, results):
            if isinstance(items, Exception):
                logger.warning("足迹搜索失败 %s: %s", q, items)
                continue
            for it in items:
                dom = _domain(it.url or "")
                if not dom or dom in seen or dom == your_domain:
                    continue
                seen.add(dom)
                out.append({"domain": dom, "url": it.url, "title": it.title or "", "hint": hint})
                if len(out) >= cap:
                    return out
        return out

    async def _emails_for(self, home_url: str, home_html: str) -> tuple[list[ProspectEmail], bool]:
        """首页 + contact/about 页抓邮箱（contact 页=高，首页=中），并判是否有表单。"""
        conf: dict[str, str] = {}
        for e in extract_emails(home_html):
            conf.setdefault(e, "中")
        has_form = has_contact_form(home_html)
        for cl in find_contact_links(home_html, home_url):
            try:
                page = await self.fetcher.fetch(cl)
            except Exception:
                continue
            html = page.raw_html or ""
            has_form = has_form or has_contact_form(html)
            for e in extract_emails(html):
                conf[e] = "高"  # contact/about 页命中，升级为高置信
        emails = [ProspectEmail(address=a, confidence=c) for a, c in conf.items()][:4]
        emails.sort(key=lambda x: {"高": 0, "中": 1, "低": 2}.get(x.confidence, 3))
        return emails, has_form

    async def _process(self, seed: dict, req: OutreachRequest, sem: asyncio.Semaphore) -> Prospect:
        async with sem:
            try:
                home = await self.fetcher.fetch(seed["url"])
                if looks_blocked(home.text):
                    raise ValueError("被反爬")
            except Exception as exc:
                logger.info("候选站抓取失败 %s: %s", seed["url"], exc)
                return Prospect(domain=seed["domain"], url=seed["url"], title=seed["title"],
                                fetched=False)
            emails, has_form = await self._emails_for(seed["url"], home.raw_html or "")
            try:
                raw = await self.llm.complete_json(
                    CLASSIFY_SYSTEM,
                    build_classify_user(req.keyword, req.your_url, req.your_brief,
                                        req.generate_emails and bool(emails),
                                        home.title or seed["title"], home.text),
                    mock=_CLS_MOCK, model=self.s.writer_model or None)
            except Exception as exc:
                logger.warning("候选站分析失败 %s: %s", seed["domain"], exc)
                raw = {}
            if not isinstance(raw, dict):
                raw = {}
            return Prospect(
                domain=seed["domain"], url=seed["url"],
                title=home.title or seed["title"],
                site_type=raw.get("site_type", ""),
                relevance=int(raw.get("relevance", 0) or 0),
                opportunity=raw.get("opportunity", "") or seed["hint"],
                reason=raw.get("reason", ""),
                emails=emails, has_form=has_form,
                email_subject=raw.get("email_subject", "") or "",
                email_body=raw.get("email_body", "") or "",
            )

    async def stream(self, req: OutreachRequest):
        """SSE 生成器：start → discovered → prospect* → done。"""
        keyword = req.keyword.strip()
        if not keyword and req.your_url:
            keyword = await self._derive_keyword(req.your_url)
            req.keyword = keyword
        if not keyword:
            yield {"type": "error", "message": "请填写主题关键词（或提供可抓取的目标网址）。"}
            return

        your_domain = _domain(req.your_url) if req.your_url else ""
        cap = min(req.max_prospects or self.s.outreach_max_prospects, self.s.outreach_max_prospects)
        yield {"type": "start", "keyword": keyword}
        seeds = await self.discover(keyword, req.location_code, req.language_code, your_domain, cap)
        if not seeds:
            yield {"type": "error", "message": "没找到候选站点。换个更常见的英文主题词试试。"}
            return
        yield {"type": "discovered", "total": len(seeds)}

        sem = asyncio.Semaphore(self.s.outreach_concurrency)
        tasks = [asyncio.create_task(self._process(s, req, sem)) for s in seeds]
        prospects: list[Prospect] = []
        done = 0
        for fut in asyncio.as_completed(tasks):
            p = await fut
            done += 1
            prospects.append(p)
            yield {"type": "prospect", "done": done, "total": len(seeds), "data": p.model_dump()}

        import base64
        xlsx = build_xlsx(prospects)
        yield {"type": "done", "total": len(seeds),
               "with_email": sum(1 for p in prospects if p.emails),
               "filename": f"外链拓客_{keyword[:20]}_{datetime.now():%Y%m%d_%H%M}.xlsx".replace("/", "_"),
               "xlsx_b64": base64.b64encode(xlsx).decode("ascii")}


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #
_HEADERS = ["域名", "页面", "站点类型", "相关度", "机会", "说明", "邮箱", "仅联系表单",
            "邮件主题", "邮件正文"]
_WIDTHS = [22, 34, 12, 8, 12, 30, 30, 10, 34, 60]


def _emails_cell(p: Prospect) -> str:
    return "\n".join(f"{e.address}（{e.confidence}）" for e in p.emails)


def build_xlsx(prospects: list[Prospect]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "外链拓客"
    ws.append(_HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True)
    for i, w in enumerate(_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # 有邮箱且非"跳过"的排前面，再按相关度
    ranked = sorted(prospects, key=lambda p: (0 if (p.emails and p.opportunity != "跳过") else 1,
                                              -p.relevance))
    wrap = Alignment(wrap_text=True, vertical="top")
    for p in ranked:
        ws.append([p.domain, p.url, p.site_type, p.relevance, p.opportunity, p.reason,
                   _emails_cell(p), "是" if (p.has_form and not p.emails) else "",
                   p.email_subject, p.email_body])
        for c in ws[ws.max_row]:
            c.alignment = wrap
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
