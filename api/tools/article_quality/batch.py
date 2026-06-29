"""Notion 导出（Markdown & CSV）→ 批量文章质量评审 → Excel。

Notion 把一个数据库导出为「Markdown & CSV」时，得到一个 ZIP：
  - 顶层一个 `<库名> <hash>.csv`（表格列，无正文）
  - 一个同名文件夹，内含每行对应的 `<标题> <hash>.md`（**完整正文**）

我们只认 .md：每个 .md = 一篇文章，首行 `# 标题` 即标题，其余为正文。
逐篇跑 ArticleAnalyzer，汇总成一个带 AI 意见的 Excel。
"""

from __future__ import annotations

import io
import re
import zipfile
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .models import ArticleCheck

# 文件名尾部 32 位 hex 是 Notion 的页面 ID，去掉只留标题
_NOTION_HASH = re.compile(r"\s+[0-9a-f]{32}(?=\.|$)")
# 导出 md 顶部的属性块：形如 `属性名: 值`（在正文之前的连续若干行）
_PROP_LINE = re.compile(r"^[^\n:]{1,40}:\s+\S")


def _clean_name(name: str) -> str:
    base = name.rsplit("/", 1)[-1]
    base = base[:-3] if base.endswith(".md") else base
    return _NOTION_HASH.sub("", base).strip()


def _strip_property_block(body: str) -> str:
    """去掉 Notion 在标题下方导出的属性块（连续的 `键: 值` 行），保留真正正文。"""
    lines = body.split("\n")
    i = 0
    while i < len(lines) and not lines[i].strip():  # 跳过前导空行
        i += 1
    j = i
    while j < len(lines) and lines[j].strip() and _PROP_LINE.match(lines[j].strip()):
        j += 1
    # 仅当属性块后面还有空行+正文时才剥离，避免误删开头就是冒号句的正文
    if j > i and j < len(lines):
        return "\n".join(lines[j:]).strip()
    return body.strip()


def parse_notion_zip(data: bytes, max_articles: int = 80) -> list[dict]:
    """解析 Notion 导出 ZIP，返回 [{filename, title, text}]（text 含 `# 标题` 供分析器识别标题）。"""
    try:
        zf = zipfile.ZipFile(io.BytesIO(data))
    except zipfile.BadZipFile as exc:
        raise ValueError("上传的不是有效的 ZIP 文件。请在 Notion 用「Export → Markdown & CSV」导出。") from exc

    out: list[dict] = []
    for info in zf.infolist():
        if info.is_dir() or not info.filename.lower().endswith(".md"):
            continue
        try:
            raw = zf.read(info).decode("utf-8", errors="replace")
        except Exception:
            continue
        lines = raw.split("\n")
        title = ""
        body = raw
        if lines and lines[0].lstrip().startswith("#"):
            title = lines[0].lstrip("#").strip()
            body = "\n".join(lines[1:])
        else:
            title = _clean_name(info.filename)
        body = _strip_property_block(body)
        # 太短的多半是数据库说明页/空页，跳过
        if len(body.strip()) < 80:
            continue
        text = f"# {title}\n\n{body}" if title else body
        out.append({"filename": _clean_name(info.filename), "title": title, "text": text})
        if len(out) >= max_articles:
            break
    return out


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #
HEADERS = [
    "文件", "标题", "语言", "字数", "综合分", "评级",
    "可读性", "信息密度", "标题分",
    "AI 意见（总评 + 优先行动）", "信息密度点评", "注水片段",
    "标题改写建议", "难句改写",
]
# 列宽（字符数）
_WIDTHS = [22, 26, 6, 7, 8, 8, 9, 9, 8, 52, 36, 40, 40, 52]


def _join_actions(c: ArticleCheck) -> str:
    parts = [c.verdict] if c.verdict else []
    parts += [f"· {a}" for a in c.priority_actions]
    return "\n".join(parts)


def _join_watery(c: ArticleCheck) -> str:
    return "\n".join(f"「{w.quote}」{w.issue} → {w.suggestion}" for w in c.density.watery[:3])


def _join_title_sugg(c: ArticleCheck) -> str:
    return "\n".join(f"{s.title}（{s.why}）" for s in c.title.suggestions[:2])


def _join_rewrites(c: ArticleCheck) -> str:
    return "\n\n".join(f"原：{r.original}\n改：{r.rewrite}" for r in c.rewrites[:3])


def _row(filename: str, c: ArticleCheck) -> list:
    return [
        filename,
        c.detected_title or "（未识别）",
        "中文" if c.language == "zh" else "英文",
        c.readability.word_count,
        c.overall_score,
        c.grade,
        c.readability.verdict,
        c.density.score,
        c.title.score,
        _join_actions(c),
        c.density.summary,
        _join_watery(c),
        _join_title_sugg(c),
        _join_rewrites(c),
    ]


def build_xlsx(results: list[tuple[str, ArticleCheck]]) -> bytes:
    """results: [(filename, ArticleCheck)]，按综合分升序（最差在前，最该改）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "文章评审"
    ws.append(HEADERS)
    head_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = head_font
        cell.alignment = Alignment(vertical="center")
    for i, w in enumerate(_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ranked = sorted(results, key=lambda x: x[1].overall_score)
    wrap = Alignment(wrap_text=True, vertical="top")
    for filename, c in ranked:
        ws.append(_row(filename, c))
        for cell in ws[ws.max_row]:
            cell.alignment = wrap
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def default_filename() -> str:
    return f"文章批量评审_{datetime.now():%Y%m%d_%H%M}.xlsx"
